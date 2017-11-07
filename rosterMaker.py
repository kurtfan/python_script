import openpyxl, pprint

print('Opening wrokbook')
wb = openpyxl.load_workbook('2014.xlsx')
sheet = wb.get_sheet_by_name('2014ALL')
roster = {}

for row in range(2, sheet.max_row + 1):
    term        = sheet['A' + str(row)].value
    course      = sheet['B' + str(row)].value
    studnetId   = sheet['C' + str(row)].value
    lastName    = sheet['D' + str(row)].value
    firstName   = sheet['E' + str(row)].value
    nationality = sheet['F' + str(row)].value

    roster.setdefault(term, {})
    roster[term].setdefault(course, {})
    roster[term][course].setdefault(studnetId, [lastName, firstName, nationality])


print('Writing to file(s)')

for i in roster:
    yearTerm = str(i)
    #print(yearTerm)
    for j in roster[i]:
        courseName = str(j)
        #print(courseName)
        wb_w = openpyxl.Workbook(write_only = True)
        ws   = wb_w.create_sheet()
        
        ws.append(['Term','Course','StudentID','LastName','FirstName','Nationality'])
        
        for k in roster[i][j]:
            studentId   = str(k)
            lastName    = str(roster[i][j][k][0])
            firstName   = str(roster[i][j][k][1])
            nationality = str(roster[i][j][k][2])
            ws.append([yearTerm, courseName, studentId, lastName, firstName, nationality])

        
        wb_w.save(str(yearTerm) + '_' + str(courseName) + '.xlsx')
print("Job is done !!!")
